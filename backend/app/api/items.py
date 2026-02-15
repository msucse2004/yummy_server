"""품목 CRUD - ADMIN 전용. 코드는 자동 생성."""
import io
from decimal import Decimal, InvalidOperation

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import select, text
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.core.auth import require_user, require_role
from app.database import get_db
from app.models import User, Item
from app.models.user import Role
from app.schemas.item import ItemCreate, ItemUpdate, ItemResponse

router = APIRouter(prefix="/api/items", tags=["items"])
RequireAdmin = Depends(require_role(Role.ADMIN))

EXCEL_HEADERS = ["코드", "상품", "무게", "단위", "단가"]


def _generate_item_code(db: Session) -> str:
    """P00001, P00002 형식의 다음 품목 코드 생성"""
    row = db.execute(
        text(
            "SELECT COALESCE(MAX(CAST(SUBSTRING(code FROM 2) AS INTEGER)), 0) + 1 AS n "
            "FROM items WHERE code ~ '^P[0-9]+$'"
        )
    ).fetchone()
    n = row[0] if row else 1
    return f"P{n:05d}"


@router.get("", response_model=list[ItemResponse])
def list_items(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    stmt = select(Item).order_by(Item.product)
    return list(db.execute(stmt).scalars().all())


@router.post("", response_model=ItemResponse, status_code=status.HTTP_201_CREATED)
def create_item(
    data: ItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    code = _generate_item_code(db)
    item = Item(code=code, **data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.get("/export/excel")
def export_items_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """품목 목록을 Excel 파일로 다운로드"""
    stmt = select(Item).order_by(Item.product)
    items = list(db.execute(stmt).scalars().all())
    wb = Workbook()
    ws = wb.active
    ws.title = "품목"
    ws.append(EXCEL_HEADERS)
    for i in items:
        ws.append([
            i.code or "",
            i.product or "",
            float(i.weight) if i.weight is not None else "",
            i.unit or "EA",
            float(i.unit_price) if i.unit_price is not None else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items.xlsx"},
    )


@router.post("/import/excel")
def import_items_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """Excel 파일에서 품목 목록 가져오기. 코드는 기존 매칭용, 신규는 자동 생성."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일을 선택해주세요")
    try:
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=400, detail="파일에 시트가 없습니다")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0
        updated = 0
        errors = []
        for i, row in enumerate(rows):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            code_val = str(row[0]).strip() if row[0] is not None else None
            product_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            # 5열: 코드,상품,무게,단위,단가 / 4열(구형): 코드,상품,단위,단가
            if len(row) >= 5:
                weight_raw = row[2] if len(row) > 2 else None
                unit_val = str(row[3]).strip() if len(row) > 3 and row[3] is not None else "EA"
                price_raw = row[4] if len(row) > 4 else None
            else:
                weight_raw = None
                unit_val = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "EA"
                price_raw = row[3] if len(row) > 3 else None
            if not product_val:
                errors.append(f"{i + 2}행: 상품명이 없습니다")
                continue
            weight = None
            if weight_raw is not None and str(weight_raw).strip() != "":
                try:
                    weight = Decimal(str(weight_raw).strip())
                except (InvalidOperation, ValueError):
                    errors.append(f"{i + 2}행: 무게가 숫자가 아닙니다 (무시됨)")
            unit_price = None
            if price_raw is not None and str(price_raw).strip() != "":
                try:
                    unit_price = Decimal(str(price_raw).strip())
                except (InvalidOperation, ValueError):
                    errors.append(f"{i + 2}행: 단가가 숫자가 아닙니다 (무시됨)")
            unit_val = unit_val if unit_val else "EA"
            existing = None
            if code_val:
                existing = db.execute(
                    select(Item).where(Item.code == code_val)
                ).scalars().first()
            if not existing:
                existing = db.execute(
                    select(Item).where(Item.product == product_val)
                ).scalars().first()
            if existing:
                existing.product = product_val
                existing.weight = weight if weight is not None else existing.weight
                existing.unit = unit_val
                existing.unit_price = unit_price if unit_price is not None else existing.unit_price
                updated += 1
            else:
                code = _generate_item_code(db)
                item = Item(
                    code=code,
                    product=product_val,
                    weight=weight,
                    unit=unit_val,
                    unit_price=unit_price,
                )
                db.add(item)
                created += 1
        db.commit()
        msg_parts = []
        if created:
            msg_parts.append(f"{created}건 등록")
        if updated:
            msg_parts.append(f"{updated}건 수정")
        msg = ", ".join(msg_parts) + " 완료" if msg_parts else "처리할 데이터가 없습니다"
        if errors:
            msg += f" (오류 {len(errors)}건: {'; '.join(errors[:5])}{'...' if len(errors) > 5 else ''})"
        return {"ok": True, "created": created, "updated": updated, "message": msg, "errors": errors}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"파일 처리 실패: {str(e)}")


@router.get("/{item_id}", response_model=ItemResponse)
def get_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    item = db.get(Item, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다")
    return item


@router.patch("/{item_id}", response_model=ItemResponse)
def update_item(
    item_id: int,
    data: ItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    item = db.get(Item, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(item, k, v)
    db.commit()
    db.refresh(item)
    return item


@router.delete("/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_item(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    item = db.get(Item, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="품목을 찾을 수 없습니다")
    db.delete(item)
    db.commit()
