"""사용자 관리 - ADMIN 전용 (기사 생성 등)"""
import io
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, field_validator
from sqlalchemy import select
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.core.auth import require_user, require_role
from app.core.security import hash_password
from app.database import get_db
from app.models import User
from app.models.user import Role
from app.schemas.auth import UserResponse

router = APIRouter(prefix="/api/users", tags=["users"])
RequireAdmin = Depends(require_role(Role.ADMIN))

EXCEL_HEADERS = ["아이디", "권한", "이름", "주민번호", "전화번호", "이력서", "상태"]
STATUS_VALUES = frozenset({"승인요청중", "재직", "퇴사"})


class UserCreateSchema(BaseModel):
    username: str = Field(..., min_length=1, max_length=64, description="한글, 영문, 숫자 등 다국어 문자 지원")
    password: str = Field(..., min_length=6)
    role: Role = Role.DRIVER
    display_name: str | None = Field(None, max_length=128)
    ssn: str | None = Field(None, max_length=32)
    phone: str | None = Field(None, max_length=32)
    resume: str | None = None
    department: str | None = Field(None, max_length=64)
    status: str | None = Field(None, max_length=32)
    preferred_locale: str | None = Field(None, max_length=64)

    @field_validator("status")
    @classmethod
    def validate_status(cls, v: str | None) -> str | None:
        if v is None or (isinstance(v, str) and v.strip() == ""):
            return None
        s = str(v).strip()
        if s in STATUS_VALUES:
            return s
        raise ValueError("상태는 승인요청중, 재직, 퇴사 중 하나여야 합니다")

class UserUpdateSchema(BaseModel):
    display_name: str | None = Field(None, max_length=128)
    role: Role | None = None
    ssn: str | None = Field(None, max_length=32)
    phone: str | None = Field(None, max_length=32)
    resume: str | None = None
    department: str | None = Field(None, max_length=64)
    status: str | None = Field(None, max_length=32)
    preferred_locale: str | None = Field(None, max_length=64)

    @field_validator("status")
    @classmethod
    def validate_status(cls, v: str | None) -> str | None:
        if v is None or v == "" or (isinstance(v, str) and v.strip() == ""):
            return None
        s = str(v).strip()
        if s in STATUS_VALUES:
            return s
        raise ValueError("상태는 승인요청중, 재직, 퇴사 중 하나여야 합니다")

    @field_validator("role", mode="before")
    @classmethod
    def role_from_str(cls, v):
        if v is None or v == "":
            return None
        if isinstance(v, Role):
            return v
        s = str(v).upper().strip()
        if s == "ADMIN":
            return Role.ADMIN
        if s == "DRIVER":
            return Role.DRIVER
        return None


@router.get("", response_model=list[UserResponse])
def list_users(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    stmt = select(User).order_by(User.id)
    return list(db.execute(stmt).scalars().all())


@router.post("", response_model=UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    data: UserCreateSchema,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    existing = db.execute(select(User).where(User.username == data.username)).scalars().first()
    if existing:
        raise HTTPException(status_code=400, detail="사용자명이 이미 존재합니다")
    user = User(
        username=data.username,
        password_hash=hash_password(data.password),
        role=data.role,
        display_name=data.display_name,
        ssn=data.ssn,
        phone=data.phone,
        resume=data.resume,
        department=data.department,
        status=data.status,
        preferred_locale=(data.preferred_locale or "").strip() or "대한민국",
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.get("/export/excel")
def export_users_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    stmt = select(User).order_by(User.username)
    users = list(db.execute(stmt).scalars().all())
    wb = Workbook()
    ws = wb.active
    ws.title = "사용자"
    ws.append(EXCEL_HEADERS)
    for u in users:
        role_display = "관리자" if u.role == Role.ADMIN else "기사"
        ws.append([
            u.username or "",
            role_display,
            u.display_name or "",
            u.ssn or "",
            u.phone or "",
            u.resume or "",
            u.status or "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users.xlsx"},
    )


@router.post("/import/excel")
def import_users_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일을 선택해주세요")
    try:
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=400, detail="파일에 시트가 없습니다")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0
        updated = 0
        errors = []
        for i, row in enumerate(rows):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            username = str(row[0]).strip() if row[0] is not None else ""
            role_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
            display_name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            ssn = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
            phone = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None
            resume = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None
            status_raw = str(row[6]).strip() if len(row) > 6 and row[6] is not None else None
            status = status_raw if status_raw and status_raw in STATUS_VALUES else None
            if status_raw and status_raw not in STATUS_VALUES:
                errors.append(f"{i + 2}행: 상태는 승인요청중, 재직, 퇴사 중 하나여야 합니다 (무시됨)")
            if not username:
                errors.append(f"{i + 2}행: 아이디가 없습니다")
                continue
            role = Role.DRIVER
            rv = (role_val or "").strip()
            if rv.upper() in ("ADMIN",) or rv in ("관리자",):
                role = Role.ADMIN
            elif rv.upper() in ("DRIVER",) or rv in ("기사",):
                role = Role.DRIVER
            existing = db.execute(select(User).where(User.username == username)).scalars().first()
            if existing:
                existing.display_name = display_name or existing.display_name
                existing.role = role
                existing.ssn = ssn or existing.ssn
                existing.phone = phone or existing.phone
                existing.resume = resume or existing.resume
                if status is not None:
                    existing.status = status
                updated += 1
            else:
                user = User(
                    username=username,
                    password_hash=hash_password("changeme123"),
                    role=role,
                    display_name=display_name,
                    ssn=ssn,
                    phone=phone,
                    resume=resume,
                    status=status,
                )
                db.add(user)
                created += 1
        db.commit()
        msg_parts = []
        if created:
            msg_parts.append(f"{created}건 등록")
        if updated:
            msg_parts.append(f"{updated}건 수정")
        msg = ", ".join(msg_parts) + " 완료" if msg_parts else "처리할 데이터가 없습니다"
        if created:
            msg += " (신규 사용자 비밀번호: changeme123)"
        if errors:
            msg += f" (오류 {len(errors)}건)"
        return {"ok": True, "created": created, "updated": updated, "message": msg, "errors": errors}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"파일 처리 실패: {str(e)}")


class SetPasswordSchema(BaseModel):
    password: str = Field(..., min_length=4)


@router.post("/{user_id}/set-password", status_code=status.HTTP_204_NO_CONTENT)
def set_user_password(
    user_id: int,
    data: SetPasswordSchema,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """관리자가 사용자 비밀번호 직접 변경"""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    user.password_hash = hash_password(data.password)
    user.must_change_password = False
    db.commit()


@router.post("/{user_id}/set-temporary-password", status_code=status.HTTP_204_NO_CONTENT)
def set_temporary_password(
    user_id: int,
    data: SetPasswordSchema,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """관리자가 임시 비밀번호 설정, 사용자가 해당 비밀번호로 로그인 시 비밀번호 변경 필수"""
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    user.password_hash = hash_password(data.password)
    user.must_change_password = True
    db.commit()


@router.patch("/{user_id}", response_model=UserResponse)
def update_user(
    user_id: int,
    data: UserUpdateSchema,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    for k, v in data.model_dump(exclude_unset=True).items():
        if k == "role" and v is None:
            continue
        setattr(user, k, v)
    db.commit()
    db.refresh(user)
    return user


@router.delete("/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_user(
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    user = db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="사용자를 찾을 수 없습니다")
    if user.username == "admin":
        raise HTTPException(status_code=400, detail="admin 사용자는 삭제할 수 없습니다")
    db.delete(user)
    db.commit()
