"""거래처 CRUD - ADMIN 전용"""
import io
import re
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import delete, select
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from app.core.auth import require_user, require_role
from app.core.security import verify_password
from app.database import get_db
from app.models import (
    AppSetting,
    Customer,
    Item,
    Photo,
    Plan,
    Route,
    RouteAssignment,
    Stop,
    StopCompletion,
    StopOrderItem,
    User,
)
from app.config import get_settings
from app.models.user import Role
from app.schemas.customer import CustomerCreate, CustomerUpdate, CustomerResponse
from app.services.contract_match import (
    contract_items_to_display_string,
    match_contract_content,
)
from app.services.geocode import maybe_geocode_and_update


class DeleteAllRequest(BaseModel):
    password: str

router = APIRouter(prefix="/api/customers", tags=["customers"])
RequireAdmin = Depends(require_role(Role.ADMIN))

EXCEL_HEADERS = ["코드", "루트", "이름", "사업자번호", "대표", "계약", "업태", "종목", "미수금액", "계약내용", "주소", "위도", "경도"]


def _get_delivery_route_count(db: Session) -> int:
    """설정의 배달 루트 개수"""
    row = db.execute(
        select(AppSetting).where(AppSetting.key == "delivery_route_count")
    ).scalars().first()
    if not row:
        return 5
    try:
        return int(row.value)
    except (ValueError, TypeError):
        return 5


def _normalize_route(route_raw: str | None, max_routes: int) -> str | None:
    """루트 텍스트를 N호차 형식으로 정규화. "1", "1호" -> "1호차" """
    if not route_raw or not str(route_raw).strip():
        return None
    s = str(route_raw).strip()
    m = re.match(r"^(\d+)", s)
    if not m:
        return s if s.endswith("호차") and s[:-2].isdigit() else None
    n = int(m.group(1))
    if 1 <= n <= max_routes:
        return f"{n}호차"
    return f"{n}호차" if n > 0 else None


def _parse_float(val: object) -> float | None:
    """Parse float from cell value, return None if invalid"""
    if val is None or str(val).strip() == "":
        return None
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return None


def _generate_customer_code(db: Session, exclude: set[str] | None = None) -> str:
    """다음 사용 가능한 거래처 코드 생성 (C0001, C0002, ...)"""
    exclude = exclude or set()
    stmt = select(Customer.code).where(Customer.code.isnot(None))
    existing = {r for r in db.execute(stmt).scalars().all() if r[0]}
    used = existing | exclude
    n = 1
    while f"C{n:04d}" in used:
        n += 1
    return f"C{n:04d}"


@router.get("", response_model=list[CustomerResponse])
def list_customers(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    stmt = select(Customer).order_by(Customer.name)
    return list(db.execute(stmt).scalars().all())


@router.post("", response_model=CustomerResponse, status_code=status.HTTP_201_CREATED)
def create_customer(
    data: CustomerCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    dump = data.model_dump()
    if not dump.get("code") or not str(dump["code"]).strip():
        dump["code"] = _generate_customer_code(db)
    else:
        existing = db.execute(select(Customer).where(Customer.code == data.code)).scalar_one_or_none()
        if existing:
            raise HTTPException(status_code=400, detail="거래처 코드가 이미 존재합니다")
    settings = get_settings()
    lat, lon = maybe_geocode_and_update(
        dump.get("address"),
        dump.get("latitude"),
        dump.get("longitude"),
        api_key=settings.kakao_rest_api_key,
    )
    if lat is not None and lon is not None:
        dump["latitude"] = lat
        dump["longitude"] = lon
    customer = Customer(**dump)
    db.add(customer)
    db.commit()
    db.refresh(customer)
    return customer


class MatchContractRequest(BaseModel):
    text: str


@router.post("/match-contract-content")
def match_contract_content_api(
    data: MatchContractRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """텍스트 계약 내용을 품목 DB와 ML 유사도 기반으로 맵핑"""
    text = (data.text or "").strip()
    matches = match_contract_content(text, db)
    display_str = contract_items_to_display_string(matches) if matches else ""
    return {"matches": matches, "display_string": display_str}


@router.get("/export/excel")
def export_customers_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """거래처 목록을 Excel 파일로 다운로드"""
    stmt = select(Customer).order_by(Customer.name)
    customers = list(db.execute(stmt).scalars().all())
    wb = Workbook()
    ws = wb.active
    ws.title = "거래처"
    ws.append(EXCEL_HEADERS)
    for c in customers:
        ws.append([
            c.code or "",
            c.route or "",
            c.name or "",
            c.business_registration_number or "",
            c.representative_name or "",
            c.contract or "",
            c.business_type or "",
            c.business_category or "",
            int(c.arrears) if c.arrears is not None else "",
            c.contract_content or "",
            c.address or "",
            float(c.latitude) if c.latitude is not None else "",
            float(c.longitude) if c.longitude is not None else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=customers.xlsx"},
    )


@router.post("/import/excel")
def import_customers_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """Excel 파일에서 거래처 목록 가져오기"""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="xlsx 파일을 선택해주세요")
    try:
        contents = file.file.read()
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            raise HTTPException(status_code=400, detail="파일에 시트가 없습니다")
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        created = 0
        updated = 0
        errors = []
        used_codes: set[str] = set()
        max_routes = _get_delivery_route_count(db)
        for i, row in enumerate(rows):
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            code_val = str(row[0]).strip() if row[0] is not None else None
            route_raw = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
            route = _normalize_route(route_raw, max_routes) if route_raw else None
            name = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            # 13열(신규): 코드,루트,이름,사업자번호,대표,계약,업태,종목,미수금액,계약내용,주소,위도,경도
            # 12열(구): 코드,루트,이름,사업자번호,대표,계약,업태,종목,계약내용,주소,위도,경도
            # 9열(구형): 코드,루트,이름,연락처,계약,계약내용,주소,위도,경도
            arrears_val = None
            if len(row) >= 13:
                business_registration_number = str(row[3]).strip() if row[3] is not None else None
                representative_name = str(row[4]).strip() if row[4] is not None else None
                contract_val = str(row[5]).strip() if row[5] is not None else None
                business_type = str(row[6]).strip() if row[6] is not None else None
                business_category = str(row[7]).strip() if row[7] is not None else None
                arrears_val = _parse_float(row[8])
                contract_content_raw = str(row[9]).strip() if row[9] is not None else None
                address = str(row[10]).strip() if row[10] is not None else None
                latitude = _parse_float(row[11])
                longitude = _parse_float(row[12])
                phone = None
            elif len(row) >= 12:
                business_registration_number = str(row[3]).strip() if row[3] is not None else None
                representative_name = str(row[4]).strip() if row[4] is not None else None
                contract_val = str(row[5]).strip() if row[5] is not None else None
                business_type = str(row[6]).strip() if row[6] is not None else None
                business_category = str(row[7]).strip() if row[7] is not None else None
                contract_content_raw = str(row[8]).strip() if row[8] is not None else None
                address = str(row[9]).strip() if row[9] is not None else None
                latitude = _parse_float(row[10])
                longitude = _parse_float(row[11])
                phone = None
            elif len(row) >= 9:
                business_registration_number = None
                representative_name = None
                business_type = None
                business_category = None
                phone = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
                contract_val = str(row[4]).strip() if len(row) > 4 and row[4] is not None else None
                contract_content_raw = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None
                address = str(row[6]).strip() if len(row) > 6 and row[6] is not None else None
                latitude = _parse_float(row[7]) if len(row) > 7 else None
                longitude = _parse_float(row[8]) if len(row) > 8 else None
            if contract_content_raw:
                matches = match_contract_content(contract_content_raw, db)
                contract_content = contract_items_to_display_string(matches) if matches else contract_content_raw
            else:
                contract_content = None
            if contract_val and contract_val not in ("계약", "해지"):
                errors.append(f"{i + 2}행: 계약은 '계약' 또는 '해지'만 가능합니다")
                continue
            contract = contract_val if contract_val in ("계약", "해지") else None
            if not name:
                errors.append(f"{i + 2}행: 이름이 없습니다")
                continue
            existing = db.execute(
                select(Customer).where(Customer.name == name)
            ).scalars().first()
            if not existing and code_val:
                existing = db.execute(
                    select(Customer).where(Customer.code == code_val)
                ).scalars().first()
            if existing:
                if code_val:
                    other = db.execute(
                        select(Customer).where(Customer.code == code_val, Customer.id != existing.id)
                    ).scalars().first()
                    if other:
                        errors.append(f"{i + 2}행: 코드 '{code_val}'가 다른 거래처에 이미 사용 중")
                        continue
                    existing.code = code_val
                existing.route = route or existing.route
                existing.name = name or existing.name
                if phone is not None:
                    existing.phone = phone or existing.phone
                existing.contract = contract or existing.contract
                if business_registration_number is not None:
                    existing.business_registration_number = business_registration_number or None
                if representative_name is not None:
                    existing.representative_name = representative_name or None
                if business_type is not None:
                    existing.business_type = business_type or None
                if business_category is not None:
                    existing.business_category = business_category or None
                if arrears_val is not None:
                    existing.arrears = arrears_val
                if contract_content is not None:
                    existing.contract_content = contract_content or None
                existing.address = address or existing.address
                if latitude is not None:
                    existing.latitude = latitude
                if longitude is not None:
                    existing.longitude = longitude
                need_geocode = (address or existing.address) and latitude is None and longitude is None
                if need_geocode:
                    _settings = get_settings()
                    geocode_addr = address if address else existing.address
                    lat, lon = maybe_geocode_and_update(
                        geocode_addr,
                        None,
                        None,
                        api_key=_settings.kakao_rest_api_key,
                    )
                    if lat is not None and lon is not None:
                        existing.latitude = lat
                        existing.longitude = lon
                updated += 1
                continue
            code = code_val if code_val else _generate_customer_code(db, exclude=used_codes)
            if not code_val:
                used_codes.add(code)
            if address and latitude is None and longitude is None:
                _settings = get_settings()
                lat, lon = maybe_geocode_and_update(
                    address, None, None, api_key=_settings.kakao_rest_api_key
                )
                if lat is not None and lon is not None:
                    latitude, longitude = lat, lon
            customer = Customer(
                code=code,
                latitude=latitude,
                longitude=longitude,
                route=route or None,
                name=name,
                phone=phone or None,
                contract=contract,
                contract_content=contract_content or None,
                address=address or None,
                memo=None,
                business_registration_number=business_registration_number or None,
                representative_name=representative_name or None,
                business_type=business_type or None,
                business_category=business_category or None,
                arrears=arrears_val,
            )
            db.add(customer)
            created += 1
        db.commit()
        msg_parts = []
        if created:
            msg_parts.append(f"{created}건 등록")
        if updated:
            msg_parts.append(f"{updated}건 수정")
        msg = ", ".join(msg_parts) + " 완료" if msg_parts else "처리할 데이터가 없습니다"
        if errors:
            msg += f" (오류 {len(errors)}건: {'; '.join(errors[:5])}{'...' if len(errors) > 5 else ''})"
        return {"ok": True, "created": created, "updated": updated, "message": msg, "errors": errors}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"파일 처리 실패: {str(e)}")


@router.post("/delete-all")
def delete_all_customers(
    data: DeleteAllRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    """거래처 전체 삭제 - admin 비밀번호 확인 후 모든 DB 데이터 삭제"""
    admin = db.execute(select(User).where(User.username == "admin")).scalars().first()
    if not admin or not verify_password(data.password, admin.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="관리자 비밀번호가 올바르지 않습니다")
    try:
        db.execute(delete(Photo))
        db.execute(delete(StopCompletion))
        db.execute(delete(StopOrderItem))
        db.execute(delete(Stop))
        db.execute(delete(RouteAssignment))
        db.execute(delete(Route))
        db.execute(delete(Plan))
        db.execute(delete(Customer))
        db.execute(delete(Item))
        db.commit()
        return {"ok": True, "message": "모든 데이터가 삭제되었습니다"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"삭제 실패: {str(e)}")


@router.get("/{customer_id}", response_model=CustomerResponse)
def get_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    customer = db.get(Customer, customer_id)
    if not customer:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    return customer


@router.patch("/{customer_id}", response_model=CustomerResponse)
def update_customer(
    customer_id: int,
    data: CustomerUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    customer = db.get(Customer, customer_id)
    if not customer:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    for k, v in data.model_dump(exclude_unset=True).items():
        if k not in ("name", "code"):
            setattr(customer, k, v)
    if customer.address and (customer.latitude is None or customer.longitude is None):
        settings = get_settings()
        lat, lon = maybe_geocode_and_update(
            customer.address,
            customer.latitude,
            customer.longitude,
            api_key=settings.kakao_rest_api_key,
        )
        if lat is not None and lon is not None:
            customer.latitude = lat
            customer.longitude = lon
    db.commit()
    db.refresh(customer)
    return customer


@router.delete("/{customer_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_user),
    _: User = RequireAdmin,
):
    customer = db.get(Customer, customer_id)
    if not customer:
        raise HTTPException(status_code=404, detail="거래처를 찾을 수 없습니다")
    db.delete(customer)
    db.commit()
